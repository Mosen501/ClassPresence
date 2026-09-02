from __future__ import annotations

import tempfile
import unittest
from types import SimpleNamespace

from openpyxl import load_workbook

from attendance_app.database import AttendanceRepository
from attendance_app.location_diagnostics import (
    analyze_classroom_reference,
    browser_family,
    summarize_location_events,
)
from attendance_app.location_reports import build_location_diagnostics_xlsx
from attendance_app.services import record_location_attempt


class LocationDiagnosticsTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.repo = AttendanceRepository(f"{self.temp_dir.name}/attendance.db")
        self.repo.init_schema()
        self.created_at = "2026-09-01T09:00:00+03:00"
        self.repo.create_course(
            code="LOC101",
            title="Location Diagnostics",
            start_date="2026-09-01",
            end_date="2026-12-01",
            total_meetings=10,
            latitude=24.8,
            longitude=46.7,
            radius_m=50.0,
            absence_limit_pct=20.0,
            created_at=self.created_at,
        )
        course = self.repo.get_course_by_code("LOC101")
        assert course is not None
        self.course_id = int(course["id"])
        self.repo.add_student_to_course(
            course_id=self.course_id,
            full_name="Location Student",
            university_id="LOC-STUDENT",
            email="location@example.edu",
            phone="",
            created_at=self.created_at,
        )
        student = self.repo.get_student_for_course(self.course_id, "LOC-STUDENT")
        assert student is not None
        self.student_id = int(student["id"])
        self.repo.add_schedule(
            course_id=self.course_id,
            weekday=1,
            label="L1",
            start_time="09:00",
            end_time="10:00",
            created_at=self.created_at,
        )
        self.schedule_id = int(
            self.repo.list_schedules_for_course(self.course_id)[0]["id"]
        )

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_failed_attempt_is_recovered_and_coordinates_are_later_redacted(self) -> None:
        self._event(
            outcome="error",
            reason_code="permission_denied",
            created_at="2026-09-01T09:01:00+03:00",
        )
        self._event(
            outcome="accepted",
            reason_code="attendance_recorded",
            created_at="2026-09-01T09:02:00+03:00",
        )

        rows = self.repo.list_location_attempt_events(course_id=self.course_id)
        failure = next(row for row in rows if row["reason_code"] == "permission_denied")
        self.assertEqual(failure["recovered_at"], "2026-09-01T09:02:00+03:00")

        redacted = self.repo.anonymize_location_coordinates_before(
            cutoff_iso="2026-10-02T00:00:00+03:00",
            redacted_at="2026-10-02T00:00:00+03:00",
        )
        self.assertEqual(redacted, 2)
        rows = self.repo.list_location_attempt_events(course_id=self.course_id)
        self.assertTrue(all(row["latitude"] is None for row in rows))
        self.assertTrue(all(row["distance_m"] is not None for row in rows))

    def test_calibration_updates_course_and_records_audit(self) -> None:
        self.repo.apply_course_location_calibration(
            course_id=self.course_id,
            latitude=24.8002,
            longitude=46.7002,
            reading_count=5,
            median_accuracy_m=8.0,
            actor_identifier="manager",
            created_at=self.created_at,
        )

        course = self.repo.get_course(self.course_id)
        assert course is not None
        self.assertAlmostEqual(float(course["latitude"]), 24.8002)
        audit = self.repo.list_course_location_calibrations(course_id=self.course_id)
        self.assertEqual(len(audit), 1)
        self.assertEqual(audit[0]["reading_count"], 5)
        self.assertAlmostEqual(float(audit[0]["previous_latitude"]), 24.8)

    def test_reference_analysis_requires_repeated_evidence_and_flags_offset(self) -> None:
        course = self.repo.get_course(self.course_id)
        assert course is not None
        events = []
        for index in range(12):
            events.append(
                {
                    "student_id": index + 1,
                    "attendance_date": "2026-09-01" if index < 6 else "2026-09-08",
                    "schedule_id": 1 if index < 6 else 2,
                    "latitude": 24.8003,
                    "longitude": 46.7003,
                    "accuracy_m": 10.0,
                }
            )

        analysis = analyze_classroom_reference(course, events)

        self.assertEqual(analysis["status"], "review")
        self.assertEqual(analysis["sample_count"], 12)
        self.assertEqual(analysis["session_count"], 2)
        self.assertGreater(float(analysis["offset_m"]), 20.0)

    def test_summary_and_excel_export_include_operational_metrics(self) -> None:
        self._event(
            outcome="rejected",
            reason_code="outside_radius",
            created_at="2026-09-01T09:01:00+03:00",
        )
        self._event(
            outcome="accepted",
            reason_code="attendance_recorded",
            created_at="2026-09-01T09:02:00+03:00",
        )
        events = self.repo.list_location_attempt_events(course_id=self.course_id)
        summary = summarize_location_events(events)
        self.assertEqual(summary["total_attempts"], 2)
        self.assertEqual(summary["accepted"], 1)
        self.assertEqual(summary["reason_counts"]["outside_radius"], 1)
        course = self.repo.get_course(self.course_id)
        assert course is not None
        report = build_location_diagnostics_xlsx(
            course=course,
            events=events,
            calibrations=[],
            reference_analysis=analyze_classroom_reference(course, events),
            generated_at=self.created_at,
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as output:
            output.write(report)
            output.flush()
            workbook = load_workbook(output.name, read_only=True)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Summary",
                "Failure Reasons",
                "Lecture Analytics",
                "Attempt Log",
                "Calibration Audit",
            ],
        )

    def test_browser_family_uses_privacy_limited_category(self) -> None:
        self.assertEqual(browser_family("Mozilla/5.0 Chrome/151 Safari/537"), "Chrome")
        self.assertEqual(browser_family("Mozilla/5.0 Version/18 Safari/605"), "Safari")

    def test_structured_permission_error_is_recorded_without_raw_user_agent(self) -> None:
        course = self.repo.get_course(self.course_id)
        student = self.repo.get_student(self.student_id)
        assert course is not None
        assert student is not None
        record_location_attempt(
            self.repo,
            SimpleNamespace(app_timezone="Asia/Riyadh"),
            course=course,
            student=student,
            geolocation_payload={
                "error": "Location access was denied.",
                "error_code": "permission_denied",
                "captured_at": self.created_at,
                "platform": "iPhone",
                "user_agent": "Mozilla/5.0 Version/18 Safari/605",
            },
            attempt_type="registration",
            success=False,
            message="Location access was denied.",
        )

        row = self.repo.list_location_attempt_events(course_id=self.course_id)[0]
        self.assertEqual(row["reason_code"], "permission_denied")
        self.assertEqual(row["browser_family"], "Safari")
        self.assertNotIn("user_agent", row)

    def _event(self, *, outcome: str, reason_code: str, created_at: str) -> None:
        self.repo.create_location_attempt_event(
            course_id=self.course_id,
            student_id=self.student_id,
            schedule_id=self.schedule_id,
            attendance_date="2026-09-01",
            attempt_type="attendance",
            outcome=outcome,
            reason_code=reason_code,
            message=reason_code,
            latitude=24.8,
            longitude=46.7,
            accuracy_m=10.0,
            distance_m=2.0,
            radius_m=50.0,
            captured_at=created_at,
            sample_count=2,
            platform="iPhone",
            browser_family="Safari",
            created_at=created_at,
        )


if __name__ == "__main__":
    unittest.main()
