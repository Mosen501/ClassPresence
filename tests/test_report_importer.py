from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook

from attendance_app.config import Settings
from attendance_app.database import AttendanceRepository
from attendance_app.report_importer import import_attendance_report_bytes


class ReportImporterTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.repo = AttendanceRepository(f"{self.temp_dir.name}/attendance.db")
        self.repo.init_schema()
        self.settings = Settings(
            app_env="development",
            app_timezone="Asia/Riyadh",
            database_target=f"{self.temp_dir.name}/attendance.db",
            manager_username="manager_user",
            manager_password_hash="unused",
            otp_delivery_mode="console",
            otp_expiry_minutes=10,
            otp_pepper="pepper",
            smtp_host="",
            smtp_port=587,
            smtp_username="",
            smtp_password="",
            smtp_sender="",
            smtp_use_tls=True,
        )

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_import_attendance_report_bytes_restores_course_data(self) -> None:
        workbook = Workbook()

        course_sheet = workbook.active
        course_sheet.title = "Course Details"
        for row in [
            ("Course Code", "MAT116"),
            ("Course Name", "Calculus"),
            ("Start Date", "2026-06-28"),
            ("End Date", "2026-08-06"),
            ("Latitude", 24.81),
            ("Longitude", 46.71),
            ("Allowed Radius (m)", 30),
            ("Absence Limit (%)", 20),
            ("Generated At", "2026-07-06T10:32:56+03:00"),
        ]:
            course_sheet.append(row)

        roster_sheet = workbook.create_sheet("Roster")
        roster_sheet.append(["Student ID", "Student Name", "Email", "Phone"])
        roster_sheet.append(["445009803", "Student One", "one@example.edu", ""])
        roster_sheet.append(["445009804", "Student Two", "two@example.edu", ""])

        timetable_sheet = workbook.create_sheet("Timetable")
        timetable_sheet.append(["Weekday", "Window Label", "Start Time", "End Time"])
        timetable_sheet.append(["Sunday", "L3", "09:20", "10:10"])
        timetable_sheet.append(["Sunday", "L4", "10:15", "11:05"])

        attendance_sheet = workbook.create_sheet("Attendance")
        attendance_sheet.append(["Student Name", "Student ID", "Date", "Window", "Stamped At", "Distance (m)"])
        attendance_sheet.append(
            ["Student One", "445009803", "2026-07-05", "L3", "2026-07-05T09:25:00+03:00", 2.5]
        )

        eligibility_sheet = workbook.create_sheet("Eligibility")
        eligibility_sheet.append(
            ["Student", "University ID", "Attended", "Absences", "Elapsed Meetings", "Total Meetings", "Threshold", "Status"]
        )
        eligibility_sheet.append(["Student One", "445009803", 1, 0, 1, 10, 2, "Eligible"])
        eligibility_sheet.append(["Student Two", "445009804", 0, 1, 1, 10, 2, "Eligible"])

        buffer = BytesIO()
        workbook.save(buffer)

        summary = import_attendance_report_bytes(
            repo=self.repo,
            settings=self.settings,
            source_name="MAT116_attendance_report3.xlsx",
            content=buffer.getvalue(),
        )

        self.assertEqual(summary["course_code"], "MAT116")
        self.assertEqual(summary["roster_rows"], 2)
        self.assertEqual(summary["schedule_rows"], 2)
        self.assertEqual(summary["imported_attendance"], 1)

        course = self.repo.get_course_by_code("MAT116")
        self.assertIsNotNone(course)
        self.assertEqual(len(self.repo.list_students_for_course(int(course["id"]))), 2)
        self.assertEqual(len(self.repo.list_schedules_for_course(int(course["id"]))), 2)
        self.assertEqual(self.repo.list_course_attendance(course_id=int(course["id"]), limit=10)[0]["university_id"], "445009803")

    def test_import_attendance_report_bytes_accepts_excel_datetime_dates(self) -> None:
        workbook = Workbook()

        course_sheet = workbook.active
        course_sheet.title = "Course Details"
        for row in [
            ("Course Code", "MAT1116"),
            ("Course Name", "Calculus 2"),
            ("Start Date", datetime(2026, 6, 28, 0, 0, 0)),
            ("End Date", datetime(2026, 8, 6, 0, 0, 0)),
            ("Latitude", 24.81),
            ("Longitude", 46.71),
            ("Allowed Radius (m)", 30),
            ("Absence Limit (%)", 20),
            ("Generated At", "2026-07-06T10:32:56+03:00"),
        ]:
            course_sheet.append(row)

        roster_sheet = workbook.create_sheet("Roster")
        roster_sheet.append(["Student ID", "Student Name", "Email", "Phone"])
        roster_sheet.append(["445009803", "Student One", "one@example.edu", ""])

        timetable_sheet = workbook.create_sheet("Timetable")
        timetable_sheet.append(["Weekday", "Window Label", "Start Time", "End Time"])
        timetable_sheet.append(["Sunday", "L3", "09:20", "10:10"])

        attendance_sheet = workbook.create_sheet("Attendance")
        attendance_sheet.append(["Student Name", "Student ID", "Date", "Window", "Stamped At", "Distance (m)"])
        attendance_sheet.append(
            ["Student One", "445009803", datetime(2026, 7, 5, 0, 0, 0), "L3", "2026-07-05T09:25:00+03:00", 2.5]
        )

        eligibility_sheet = workbook.create_sheet("Eligibility")
        eligibility_sheet.append(
            ["Student", "University ID", "Attended", "Absences", "Elapsed Meetings", "Total Meetings", "Threshold", "Status"]
        )
        eligibility_sheet.append(["Student One", "445009803", 1, 0, 1, 10, 2, "Eligible"])

        buffer = BytesIO()
        workbook.save(buffer)

        summary = import_attendance_report_bytes(
            repo=self.repo,
            settings=self.settings,
            source_name="MAT1116_attendance_report.xlsx",
            content=buffer.getvalue(),
        )

        self.assertEqual(summary["course_code"], "MAT1116")
        self.assertEqual(summary["imported_attendance"], 1)

        course = self.repo.get_course_by_code("MAT1116")
        self.assertIsNotNone(course)
        self.assertEqual(course["start_date"], "2026-06-28")
        self.assertEqual(course["end_date"], "2026-08-06")
        attendance_row = self.repo.list_course_attendance(course_id=int(course["id"]), limit=10)[0]
        self.assertEqual(attendance_row["attendance_date"], "2026-07-05")


if __name__ == "__main__":
    unittest.main()
