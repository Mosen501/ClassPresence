from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import Workbook

from attendance_app.roster import parse_roster_file


class RosterParserTestCase(unittest.TestCase):
    def test_parse_xlsx_roster(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["student id", "student name", "email"])
        worksheet.append(["U2026001", "Amina Yusuf", "amina@example.edu"])

        buffer = BytesIO()
        workbook.save(buffer)

        rows = parse_roster_file("roster.xlsx", buffer.getvalue())

        self.assertEqual(
            rows,
            [
                {
                    "university_id": "U2026001",
                    "full_name": "Amina Yusuf",
                    "email": "amina@example.edu",
                    "phone": "",
                }
            ],
        )

    def test_parse_csv_roster(self) -> None:
        content = (
            "student id,student name,email\n"
            "U2026001,Amina Yusuf,amina@example.edu\n"
            "U2026002,Daniel Okoro,daniel@example.edu\n"
        ).encode("utf-8")

        rows = parse_roster_file("roster.csv", content)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1]["university_id"], "U2026002")

    def test_reject_missing_required_column(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["student id", "student name"])
        worksheet.append(["U2026001", "Amina Yusuf"])

        buffer = BytesIO()
        workbook.save(buffer)

        with self.assertRaisesRegex(ValueError, "missing required columns: email"):
            parse_roster_file("roster.xlsx", buffer.getvalue())

    def test_reject_duplicate_student_ids(self) -> None:
        content = (
            "student id,student name,email\n"
            "U2026001,Amina Yusuf,amina@example.edu\n"
            "U2026001,Another Student,another@example.edu\n"
        ).encode("utf-8")

        with self.assertRaisesRegex(ValueError, "Duplicate student ID"):
            parse_roster_file("roster.csv", content)

    def test_parse_extensionless_xlsx_with_arabic_headers(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["اسم الطالب", "الرقم الجامعي", "email"])
        worksheet.append(["طالب تجريبي", "445009803", "445009803@example.edu"])

        buffer = BytesIO()
        workbook.save(buffer)

        rows = parse_roster_file("exportExcel", buffer.getvalue())

        self.assertEqual(
            rows,
            [
                {
                    "university_id": "445009803",
                    "full_name": "طالب تجريبي",
                    "email": "445009803@example.edu",
                    "phone": "",
                }
            ],
        )


if __name__ == "__main__":
    unittest.main()
