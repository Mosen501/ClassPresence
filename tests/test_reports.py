from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timedelta
from io import BytesIO
from zipfile import ZipFile
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from attendance_app.config import Settings
from attendance_app.database import AttendanceRepository
from attendance_app.report_importer import import_attendance_report_bytes
from attendance_app.reports import REPORT_SHEETS, build_course_report_xlsx


class ReportsTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.generated_at = datetime(2026, 8, 30, 12, 30, tzinfo=ZoneInfo("Asia/Riyadh"))
        self.course = {
            "id": 1,
            "code": "MAT1116",
            "title": "Foundations of Mathematics",
            "start_date": "2026-08-30",
            "end_date": "2026-12-01",
            "total_meetings": 24,
            "latitude": 24.7136,
            "longitude": 46.6753,
            "radius_m": 20.0,
            "absence_limit_pct": 20.0,
        }
        self.students = [
            {
                "id": 1,
                "university_id": "U2026001",
                "full_name": "Amina Yusuf",
                "email": "amina.yusuf@example.edu",
                "phone": "+1555000001",
                "registered_device_id": 41,
                "device_registered_at": "2026-08-30T08:00:00+03:00",
                "device_last_used_at": "2026-08-30T09:05:00+03:00",
                "device_type": "single_device",
                "device_backed_up": 0,
            },
            {
                "id": 2,
                "university_id": "U2026002",
                "full_name": "Daniel Okoro",
                "email": "daniel.okoro@example.edu",
                "phone": "+1555000002",
                "registered_device_id": None,
                "device_registered_at": None,
                "device_last_used_at": None,
                "device_type": None,
                "device_backed_up": None,
            },
        ]
        self.schedules = [
            {
                "id": 10,
                "weekday": 6,
                "label": "Morning Lecture",
                "start_time": "09:00",
                "end_time": "10:00",
            }
        ]
        self.attendance_records = [
            {
                "attendance_id": 91,
                "schedule_id": 10,
                "full_name": "Amina Yusuf",
                "university_id": "U2026001",
                "attendance_date": "2026-08-30",
                "schedule_label": "Morning Lecture",
                "schedule_start_time": "09:00",
                "schedule_end_time": "10:00",
                "stamped_at": "2026-08-30T09:05:00+03:00",
                "student_latitude": 24.71361,
                "student_longitude": 46.67531,
                "distance_m": 1.5,
                "accuracy_m": 4.0,
                "registered_device_id": 41,
                "device_binding_hash": "attendance-raw-binding-secret",
            }
        ]
        self.eligibility_rows = [
            {
                "Student": "Amina Yusuf",
                "University ID": "U2026001",
                "Attended": 1,
                "Absences": 0,
                "Elapsed Meetings": 1,
                "Total Meetings": 24,
                "Threshold": 5,
                "Status": "Eligible",
            },
            {
                "Student": "Daniel Okoro",
                "University ID": "U2026002",
                "Attended": 0,
                "Absences": 1,
                "Elapsed Meetings": 1,
                "Total Meetings": 24,
                "Threshold": 5,
                "Status": "Eligible",
            },
        ]
        self.security_alerts = [
            {
                "id": 7,
                "created_at": "2026-08-30T09:03:00+03:00",
                "attendance_date": "2026-08-30",
                "full_name": "Amina Yusuf",
                "university_id": "U2026001",
                "schedule_label": "Morning Lecture",
                "severity": "critical",
                "alert_type": "device_linked_to_another_student",
                "message": "Blocked device reuse attempt.",
                "device_binding_hash": "alert-raw-binding-secret",
                "latitude": 24.71361,
                "longitude": 46.67531,
                "accuracy_m": 5.0,
                "resolved_at": None,
            }
        ]
        self.device_audit_events = [
            {
                "id": 3,
                "created_at": "2026-08-30T08:00:00+03:00",
                "student_name": "Amina Yusuf",
                "university_id": "U2026001",
                "event_type": "student_device_registered",
                "actor_type": "student",
                "actor_identifier": "U2026001",
                "course_code": "MAT1116",
                "previous_device_id": None,
                "new_device_id": 41,
            }
        ]
        self.otp_activity = [
            {
                "id": 55,
                "created_at": "2026-08-30T08:58:00+03:00",
                "attendance_date": "2026-08-30",
                "full_name": "Amina Yusuf",
                "university_id": "U2026001",
                "schedule_label": "Morning Lecture",
                "delivery_method": "email",
                "delivery_target": "private.target@example.edu",
                "expires_at": "2026-08-30T09:08:00+03:00",
                "used_at": "2026-08-30T09:01:00+03:00",
                "invalidated_at": None,
            }
        ]

    def _build_report(self) -> bytes:
        return build_course_report_xlsx(
            course=self.course,
            students=self.students,
            schedules=self.schedules,
            attendance_records=self.attendance_records,
            eligibility_rows=self.eligibility_rows,
            security_alerts=self.security_alerts,
            device_audit_events=self.device_audit_events,
            otp_activity=self.otp_activity,
            generated_at=self.generated_at,
        )

    def test_complete_report_has_all_sheets_formulas_charts_and_masking(self) -> None:
        workbook = load_workbook(BytesIO(self._build_report()), data_only=False)

        self.assertEqual(workbook.sheetnames, REPORT_SHEETS)
        self.assertEqual(len(workbook["Executive Summary"]._charts), 2)
        self.assertIn("Student Performance", workbook["Executive Summary"]["E5"].value)
        self.assertTrue(str(workbook["Student Performance"]["D6"].value).startswith("=MAX"))
        self.assertTrue(str(workbook["Student Performance"]["I6"].value).startswith("=IF"))
        self.assertTrue(str(workbook["Lecture Analytics"]["I6"].value).startswith("=IFERROR"))
        self.assertIsInstance(workbook["Attendance Records"]["D6"].value, datetime)
        self.assertIsInstance(workbook["Attendance Records"]["H6"].value, datetime)
        self.assertEqual(workbook["Roster"]["F6"].value, "DEV-000041")
        self.assertEqual(workbook["Roster"]["D6"].value, "tel:+1555000001")
        self.assertEqual(workbook["OTP Activity"]["H6"].value, "p*************@example.edu")
        self.assertNotEqual(workbook["Security Alerts"]["J6"].value, "alert-raw-binding-secret")
        self.assertEqual(len(workbook["Attendance Records"].tables), 0)
        self.assertEqual(len(workbook["Security Alerts"].tables), 0)
        self.assertEqual(workbook["Attendance Records"].auto_filter.ref, "A5:V6")
        self.assertEqual(workbook["Security Alerts"].auto_filter.ref, "A5:O6")

        workbook_text = "\n".join(
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        self.assertNotIn("attendance-raw-binding-secret", workbook_text)
        self.assertNotIn("alert-raw-binding-secret", workbook_text)
        self.assertNotIn("private.target@example.edu", workbook["OTP Activity"]["H6"].value)
        self.assertNotIn("code_hash", workbook_text.lower())
        self.assertNotIn("public_key", workbook_text.lower())
        self.assertNotIn("credential_id", workbook_text.lower())

    def test_report_package_uses_filters_without_excel_table_parts(self) -> None:
        report = self._build_report()

        with ZipFile(BytesIO(report)) as archive:
            names = archive.namelist()
            self.assertFalse(any(name.startswith("xl/tables/") for name in names))
            self.assertNotIn(b"/table", archive.read("[Content_Types].xml"))
            for name in (item for item in names if item.startswith("xl/worksheets/sheet")):
                self.assertNotIn(b"<tableParts", archive.read(name))

        workbook = load_workbook(BytesIO(report), data_only=False)
        for sheet_name in REPORT_SHEETS[2:]:
            sheet = workbook[sheet_name]
            self.assertEqual(len(sheet.tables), 0)
            self.assertTrue(str(sheet.auto_filter.ref).startswith("A5:"))

    def test_report_formulas_have_no_broken_or_unsupported_references(self) -> None:
        workbook = load_workbook(BytesIO(self._build_report()), data_only=False)
        formulas = [
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        ]

        self.assertGreater(len(formulas), 0)
        for formula in formulas:
            self.assertTrue(formula.startswith("="))
            self.assertNotIn("#REF!", formula.upper())
            self.assertNotIn("#NAME?", formula.upper())
            self.assertNotIn("_XLFN.", formula.upper())
            self.assertNotIn("[", formula)

    def test_new_report_restores_operational_data_only(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            database_path = f"{temp_dir}/attendance.db"
            repo = AttendanceRepository(database_path)
            repo.init_schema()
            settings = Settings(
                app_env="development",
                app_timezone="Asia/Riyadh",
                database_target=database_path,
                manager_username="manager",
                manager_password_hash="unused",
                otp_delivery_mode="console",
                otp_expiry_minutes=10,
                otp_pepper="pepper",
                smtp_host="",
                smtp_port=587,
                smtp_username="",
                smtp_password="",
                smtp_sender="",
                smtp_use_tls=True,
            )

            result = import_attendance_report_bytes(
                repo=repo,
                settings=settings,
                source_name="MAT1116_complete_report.xlsx",
                content=self._build_report(),
            )

            course = repo.get_course_by_code("MAT1116")
            assert course is not None
            self.assertEqual(result["roster_rows"], 2)
            self.assertEqual(result["schedule_rows"], 1)
            self.assertEqual(result["imported_attendance"], 1)
            self.assertEqual(course["total_meetings"], 14)
            restored_student = repo.get_student_for_course(int(course["id"]), "U2026001")
            assert restored_student is not None
            self.assertEqual(restored_student["phone"], "+1555000001")
            self.assertEqual(repo.list_proxy_alerts_for_report(course_id=int(course["id"])), [])
            self.assertEqual(
                repo.list_device_audit_events_for_report(course_id=int(course["id"])), []
            )
            self.assertEqual(repo.list_otp_activity_for_report(course_id=int(course["id"])), [])

    def test_report_escapes_spreadsheet_formula_injection(self) -> None:
        self.students[0]["full_name"] = '=HYPERLINK("https://example.test","Open")'

        workbook = load_workbook(BytesIO(self._build_report()), data_only=False)
        student_name = workbook["Roster"]["B6"]

        self.assertEqual(student_name.data_type, "s")
        self.assertTrue(str(student_name.value).startswith("'="))
        self.assertTrue(student_name.quotePrefix)

    def test_empty_report_avoids_blank_charts_and_fake_rows(self) -> None:
        report = build_course_report_xlsx(
            course=self.course,
            students=[],
            schedules=[],
            attendance_records=[],
            eligibility_rows=[],
            generated_at=self.generated_at,
        )

        workbook = load_workbook(BytesIO(report), data_only=False)
        self.assertEqual(workbook.sheetnames, REPORT_SHEETS)
        self.assertEqual(len(workbook["Executive Summary"]._charts), 0)
        self.assertIsNone(workbook["Roster"]["A6"].value)
        self.assertIsNone(workbook["Attendance Records"]["A6"].value)

    def test_report_attendance_query_is_not_limited_to_dashboard_page_size(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            repo = AttendanceRepository(f"{temp_dir}/attendance.db")
            repo.init_schema()
            repo.create_course(
                code="SEC200",
                title="Security Reporting",
                start_date="2026-01-01",
                end_date="2027-12-31",
                total_meetings=200,
                latitude=1.0,
                longitude=1.0,
                radius_m=20,
                absence_limit_pct=20,
                created_at="2026-01-01T08:00:00+00:00",
            )
            course = repo.get_course_by_code("SEC200")
            assert course is not None
            repo.add_student_to_course(
                course_id=int(course["id"]),
                full_name="Report Student",
                university_id="R001",
                email="report@example.edu",
                phone="",
                created_at="2026-01-01T08:00:00+00:00",
            )
            student = repo.get_student_for_course(int(course["id"]), "R001")
            assert student is not None
            repo.add_schedule(
                course_id=int(course["id"]),
                weekday=3,
                label="Lecture",
                start_time="08:00",
                end_time="09:00",
                created_at="2026-01-01T08:00:00+00:00",
            )
            schedule = repo.list_schedules_for_course(int(course["id"]))[0]
            start = datetime(2026, 1, 1, 8, 5)
            for index in range(125):
                timestamp = start + timedelta(days=index)
                repo.record_attendance(
                    course_id=int(course["id"]),
                    student_id=int(student["id"]),
                    schedule_id=int(schedule["id"]),
                    attendance_date=timestamp.date().isoformat(),
                    stamped_at=timestamp.isoformat(),
                    student_latitude=1.0,
                    student_longitude=1.0,
                    accuracy_m=5,
                    distance_m=0,
                    device_info="{}",
                )

            self.assertEqual(
                len(repo.list_course_attendance_for_report(course_id=int(course["id"]))),
                125,
            )


if __name__ == "__main__":
    unittest.main()
