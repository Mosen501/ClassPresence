from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook


HEADER_ALIASES = {
    "university_id": {
        "studentid",
        "universityid",
        "studentnumber",
        "universitynumber",
        "الرقمالجامعي",
    },
    "full_name": {"studentname", "fullname", "name", "اسمالطالب"},
    "email": {"email", "emailaddress", "البريدالالكتروني"},
}


def parse_roster_file(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        rows = _parse_csv(content)
    elif suffix == ".xlsx" or _looks_like_xlsx(content):
        rows = _parse_xlsx(content)
    else:
        raise ValueError("Only .xlsx and .csv roster files are supported.")

    normalized_rows = _normalize_rows(rows)
    if not normalized_rows:
        raise ValueError("The uploaded roster file did not contain any student rows.")
    return normalized_rows


def _parse_csv(content: bytes) -> list[dict[str, object]]:
    text_stream = StringIO(content.decode("utf-8-sig"))
    reader = csv.DictReader(text_stream)
    if reader.fieldnames is None:
        raise ValueError("The CSV file is missing a header row.")
    return list(reader)


def _parse_xlsx(content: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration as error:
        raise ValueError("The Excel file is empty.") from error

    headers = [str(value).strip() if value is not None else "" for value in header_row]
    if not any(headers):
        raise ValueError("The Excel file is missing a header row.")

    rows: list[dict[str, object]] = []
    for values in iterator:
        rows.append({headers[index]: value for index, value in enumerate(values)})
    return rows


def _normalize_rows(rows: list[dict[str, object]]) -> list[dict[str, str]]:
    normalized_column_map = _build_column_map(rows[0].keys()) if rows else {}
    required_columns = {"university_id", "full_name", "email"}
    missing_columns = sorted(required_columns - set(normalized_column_map))
    if missing_columns:
        raise ValueError(
            "Roster file is missing required columns: "
            + ", ".join(_display_column_name(column) for column in missing_columns)
        )

    normalized_rows: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for row_number, raw_row in enumerate(rows, start=2):
        normalized_row = {
            "university_id": _clean_cell(raw_row.get(normalized_column_map["university_id"])),
            "full_name": _clean_cell(raw_row.get(normalized_column_map["full_name"])),
            "email": _clean_cell(raw_row.get(normalized_column_map["email"])),
            "phone": "",
        }
        if not any(normalized_row.values()):
            continue

        missing_values = [
            column
            for column in ("university_id", "full_name", "email")
            if not normalized_row[column]
        ]
        if missing_values:
            raise ValueError(
                f"Row {row_number} is missing: "
                + ", ".join(_display_column_name(column) for column in missing_values)
            )

        if normalized_row["university_id"] in seen_ids:
            raise ValueError(
                f"Duplicate student ID '{normalized_row['university_id']}' found in row {row_number}."
            )
        seen_ids.add(normalized_row["university_id"])
        normalized_rows.append(normalized_row)

    return normalized_rows


def _build_column_map(headers) -> dict[str, str]:
    column_map: dict[str, str] = {}
    for header in headers:
        normalized_header = _normalize_header(header)
        for target, aliases in HEADER_ALIASES.items():
            if normalized_header in aliases and target not in column_map:
                column_map[target] = str(header)
    return column_map


def _normalize_header(value: object) -> str:
    return "".join(character for character in str(value).strip().lower() if character.isalnum())


def _looks_like_xlsx(content: bytes) -> bool:
    return content[:4] == b"PK\x03\x04"


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _display_column_name(column: str) -> str:
    return {
        "university_id": "student id",
        "full_name": "student name",
        "email": "email",
    }[column]
