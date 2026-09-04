from __future__ import annotations

import json
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from attendance_app.database import AttendanceRepository
from attendance_app.services import now_in_app_timezone

WEEKDAY_MAP = {
    "Monday": 0,
    "Tuesday": 1,
    "Wednesday": 2,
    "Thursday": 3,
    "Friday": 4,
    "Saturday": 5,
    "Sunday": 6,
}


def import_attendance_report_bytes(
    *,
    repo: AttendanceRepository,
    settings,
    source_name: str,
    content: bytes,
) -> dict[str, int | str]:
    workbook = load_workbook(BytesIO(content), data_only=True)

    course_sheet = workbook["Course Details"]
    course_details = {
        _normalize(row[0]): row[1]
        for row in course_sheet.iter_rows(min_row=1, values_only=True)
        if row[0]
    }
    course_code = _normalize(course_details["Course Code"]).upper()
    course_title = _normalize(course_details["Course Name"])
    start_date = _normalize_iso_date(course_details["Start Date"])
    end_date = _normalize_iso_date(course_details["End Date"]) or start_date
    latitude = float(course_details["Latitude"])
    longitude = float(course_details["Longitude"])
    radius_m = float(course_details["Allowed Radius (m)"])
    absence_limit_pct = float(course_details["Absence Limit (%)"])
    generated_at = (
        _normalize_timestamp(course_details["Generated At"])
        or now_in_app_timezone(settings).isoformat()
    )

    if course_details.get("Total Meetings") not in (None, ""):
        total_meetings = max(1, int(course_details["Total Meetings"]))
    else:
        performance_sheet = _get_sheet(workbook, "Student Performance", "Eligibility")
        performance_header_row, performance_headers = _find_header_row(
            performance_sheet,
            {"total meetings"},
        )
        total_meetings_index = performance_headers["total meetings"]
        total_meetings = max(
            1,
            max(
                (
                    int(row[total_meetings_index])
                    for row in performance_sheet.iter_rows(
                        min_row=performance_header_row + 1,
                        values_only=True,
                    )
                    if row and row[total_meetings_index] is not None
                ),
                default=1,
            ),
        )

    roster_sheet = workbook["Roster"]
    roster_rows = []
    roster_header_row, roster_headers = _find_header_row(
        roster_sheet,
        {"student id", "student name", "email", "phone"},
    )
    for row in roster_sheet.iter_rows(min_row=roster_header_row + 1, values_only=True):
        student_id = row[roster_headers["student id"]]
        if student_id is None:
            continue
        roster_rows.append(
            {
                "university_id": _normalize(student_id),
                "full_name": _normalize(row[roster_headers["student name"]]),
                "email": _normalize(row[roster_headers["email"]]),
                "phone": _normalize_phone(row[roster_headers["phone"]]),
            }
        )

    timetable_sheet = workbook["Timetable"]
    schedule_rows = []
    timetable_header_row, timetable_headers = _find_header_row(
        timetable_sheet,
        {"weekday", "window label", "start time", "end time"},
    )
    for row in timetable_sheet.iter_rows(
        min_row=timetable_header_row + 1,
        values_only=True,
    ):
        weekday_name = row[timetable_headers["weekday"]]
        if weekday_name is None:
            continue
        schedule_rows.append(
            {
                "weekday": WEEKDAY_MAP[_normalize(weekday_name)],
                "label": _normalize(row[timetable_headers["window label"]]),
                "start_time": _normalize(row[timetable_headers["start time"]]),
                "end_time": _normalize(row[timetable_headers["end time"]]),
                "attendance_grace_minutes": int(
                    row[timetable_headers["on-time grace (min)"]]
                    if "on-time grace (min)" in timetable_headers
                    and row[timetable_headers["on-time grace (min)"]] not in (None, "")
                    else 10
                ),
            }
        )

    existing_course = repo.get_course_by_code(course_code)
    if existing_course is None:
        repo.create_course(
            code=course_code,
            title=course_title,
            start_date=start_date,
            end_date=end_date,
            total_meetings=total_meetings,
            latitude=latitude,
            longitude=longitude,
            radius_m=radius_m,
            absence_limit_pct=absence_limit_pct,
            created_at=generated_at,
        )
    else:
        repo.update_course(
            course_id=int(existing_course["id"]),
            code=course_code,
            title=course_title,
            start_date=start_date,
            end_date=end_date,
            latitude=latitude,
            longitude=longitude,
            radius_m=radius_m,
            absence_limit_pct=absence_limit_pct,
            actor_identifier=f"report import: {source_name}",
            updated_at=generated_at,
        )

    course = repo.get_course_by_code(course_code)
    if course is None:
        raise RuntimeError("Course import failed.")
    course_id = int(course["id"])

    repo.sync_course_roster(
        course_id=course_id,
        roster_rows=roster_rows,
        created_at=generated_at,
    )
    repo.sync_course_schedules(
        course_id=course_id,
        schedule_rows=schedule_rows,
        created_at=generated_at,
    )
    repo.update_course_total_meetings(
        course_id=course_id,
        total_meetings=total_meetings,
    )

    students_by_university_id = {
        str(row["university_id"]): row for row in repo.list_students_for_course(course_id)
    }
    schedules_by_key = {
        (int(row["weekday"]), str(row["label"])): row
        for row in repo.list_schedules_for_course(course_id)
    }

    attendance_sheet = _get_sheet(workbook, "Attendance Records", "Attendance")
    attendance_header_row, attendance_headers = _find_header_row(
        attendance_sheet,
        {"student id", "date", "distance (m)"},
    )
    lecture_header = "lecture" if "lecture" in attendance_headers else "window"
    timestamp_header = "checked in" if "checked in" in attendance_headers else "stamped at"
    if lecture_header not in attendance_headers or timestamp_header not in attendance_headers:
        raise ValueError("The attendance sheet is missing lecture or timestamp columns.")
    imported_attendance = 0
    skipped_attendance = 0
    placeholder_device_info = json.dumps(
        {
            "imported_from_report": True,
            "source_file": source_name,
            "original_student_coordinates_unavailable": True,
        },
        ensure_ascii=False,
    )

    for row in attendance_sheet.iter_rows(
        min_row=attendance_header_row + 1,
        values_only=True,
    ):
        student_id = row[attendance_headers["student id"]]
        if student_id is None:
            continue
        attendance_date = row[attendance_headers["date"]]
        window_label = row[attendance_headers[lecture_header]]
        stamped_at = row[attendance_headers[timestamp_header]]
        distance_m = row[attendance_headers["distance (m)"]]
        attendance_status = (
            _normalize(row[attendance_headers["attendance status"]]).lower()
            if "attendance status" in attendance_headers
            else "present"
        )
        if attendance_status not in {
            "present",
            "late",
            "instructor_present",
            "instructor_late",
        }:
            attendance_status = "present"
        record_source = (
            _normalize(row[attendance_headers["record source"]]).lower()
            if "record source" in attendance_headers
            else "import"
        )
        override_reason = (
            _normalize(row[attendance_headers["exception reason"]])
            if "exception reason" in attendance_headers
            else ""
        )
        recorded_by = (
            _normalize(row[attendance_headers["recorded by"]])
            if "recorded by" in attendance_headers
            else ""
        )

        university_id = _normalize(student_id)
        student = students_by_university_id.get(university_id)
        if student is None:
            skipped_attendance += 1
            continue

        attendance_day = _coerce_date(attendance_date)
        schedule = schedules_by_key.get((attendance_day.weekday(), _normalize(window_label)))
        if schedule is None:
            skipped_attendance += 1
            continue

        if repo.attendance_exists(
            course_id=course_id,
            student_id=int(student["id"]),
            schedule_id=int(schedule["id"]),
            attendance_date=attendance_day.isoformat(),
        ):
            skipped_attendance += 1
            continue

        repo.record_attendance(
            course_id=course_id,
            student_id=int(student["id"]),
            schedule_id=int(schedule["id"]),
            attendance_date=attendance_day.isoformat(),
            stamped_at=_normalize_timestamp(stamped_at),
            student_latitude=latitude,
            student_longitude=longitude,
            accuracy_m=None,
            distance_m=float(distance_m or 0.0),
            device_info=placeholder_device_info,
            schedule_label_snapshot=str(schedule["label"]),
            schedule_start_time_snapshot=str(schedule["start_time"]),
            schedule_end_time_snapshot=str(schedule["end_time"]),
            reference_latitude=latitude,
            reference_longitude=longitude,
            reference_radius_m=radius_m,
            attendance_status=attendance_status,
            record_source=record_source or "import",
            override_reason=override_reason,
            recorded_by=recorded_by,
            evidence_snapshot_source="report_import",
        )
        imported_attendance += 1

    return {
        "course_code": course_code,
        "course_id": course_id,
        "roster_rows": len(roster_rows),
        "schedule_rows": len(schedule_rows),
        "imported_attendance": imported_attendance,
        "skipped_attendance": skipped_attendance,
    }


def _normalize(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.startswith("'") and text[1:2] in {"=", "+", "-", "@"}:
        return text[1:]
    return text


def _coerce_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _normalize(value)
    if not text:
        raise ValueError("A required date value is blank in the attendance report.")

    try:
        return date.fromisoformat(text)
    except ValueError:
        try:
            return datetime.fromisoformat(text).date()
        except ValueError as error:
            raise ValueError(f"Unsupported date value in the attendance report: {text}") from error


def _normalize_iso_date(value) -> str:
    if value is None:
        return ""
    return _coerce_date(value).isoformat()


def _normalize_timestamp(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return _normalize(value)


def _get_sheet(workbook, *names: str):
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    raise ValueError(f"The workbook is missing a required sheet: {' or '.join(names)}.")


def _find_header_row(sheet, required_headers: set[str]) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True),
        start=1,
    ):
        headers = {
            _normalize_header(value): index
            for index, value in enumerate(row)
            if value not in (None, "")
        }
        if required_headers.issubset(headers):
            return row_number, headers
    raise ValueError(f"The {sheet.title} sheet is missing required report columns.")


def _normalize_header(value) -> str:
    return " ".join(_normalize(value).lower().split())


def _normalize_phone(value) -> str:
    text = _normalize(value)
    return text[4:] if text.lower().startswith("tel:") else text
