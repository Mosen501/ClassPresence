from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from attendance_app.location_diagnostics import (
    LOCATION_REASON_LABELS,
    build_lecture_location_summary,
    summarize_location_events,
)


HEADER_FILL = PatternFill("solid", fgColor="132238")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_location_diagnostics_xlsx(
    *,
    course: dict,
    events: list[dict],
    calibrations: list[dict],
    location_changes: list[dict] | None = None,
    reference_analysis: dict,
    generated_at: str,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "ClassPresence"
    workbook.properties.title = f"{course['code']} location diagnostics"

    summary = summarize_location_events(events)
    summary_sheet = workbook.create_sheet("Summary")
    _write_rows(
        summary_sheet,
        ["Metric", "Value"],
        [
            ["Course", course["code"]],
            ["Generated", generated_at],
            ["Configured latitude", course["latitude"]],
            ["Configured longitude", course["longitude"]],
            ["Configured radius (m)", course["radius_m"]],
            ["Total attempts", summary["total_attempts"]],
            ["Unique students", summary["unique_students"]],
            ["Accepted", summary["accepted"]],
            ["Location failures", summary["failures"]],
            ["Attempt success rate", summary["success_rate"] / 100],
            ["Attendance windows attempted", summary["attendance_windows_attempted"]],
            ["Attendance windows completed", summary["attendance_windows_completed"]],
            ["Attendance completion rate", summary["attendance_completion_rate"] / 100],
            ["Students with unresolved failures", summary["students_with_unresolved_failures"]],
            ["Median attempts per window", summary["median_attempts_per_window"]],
            ["Recovered failures", summary["recovered_failures"]],
            ["Reference status", reference_analysis["status"]],
            ["Reference offset (m)", reference_analysis.get("offset_m")],
            ["Reference sample count", reference_analysis["sample_count"]],
            ["Reference lecture count", reference_analysis["session_count"]],
            ["Reference analysis", reference_analysis["message"]],
        ],
    )
    summary_sheet["B11"].number_format = "0.0%"
    summary_sheet["B14"].number_format = "0.0%"

    reason_sheet = workbook.create_sheet("Failure Reasons")
    reason_rows = []
    for reason, count in sorted(
        summary["reason_counts"].items(), key=lambda item: (-item[1], item[0])
    ):
        reason_rows.append(
            [
                LOCATION_REASON_LABELS.get(reason, reason.replace("_", " ").title()),
                reason,
                count,
                summary["affected_students"].get(reason, 0),
            ]
        )
    _write_rows(
        reason_sheet,
        ["Reason", "Code", "Attempts", "Affected students"],
        reason_rows,
    )

    lecture_sheet = workbook.create_sheet("Lecture Analytics")
    lecture_rows = [
        [
            row["attendance_date"],
            row["schedule_label"],
            row["total_attempts"],
            row["unique_students"],
            row["accepted"],
            row["reason_counts"].get("outside_radius", 0),
            row["reason_counts"].get("poor_accuracy", 0),
            row["reason_counts"].get("permission_denied", 0),
            row["reason_counts"].get("timeout", 0),
            row["attendance_windows_attempted"],
            row["attendance_windows_completed"],
            row["attendance_completion_rate"] / 100,
        ]
        for row in build_lecture_location_summary(events)
    ]
    _write_rows(
        lecture_sheet,
        [
            "Date",
            "Window",
            "Attempts",
            "Students",
            "Accepted",
            "Outside radius",
            "Poor accuracy",
            "Permission denied",
            "Timeout",
            "Attendance windows attempted",
            "Attendance windows completed",
            "Attendance completion rate",
        ],
        lecture_rows,
    )
    for cell in lecture_sheet["L"][1:]:
        cell.number_format = "0.0%"

    attempts_sheet = workbook.create_sheet("Attempt Log")
    _write_rows(
        attempts_sheet,
        [
            "Time",
            "Date",
            "Window",
            "Evidence snapshot",
            "Student",
            "Student ID",
            "Type",
            "Outcome",
            "Reason",
            "Accuracy (m)",
            "Distance (m)",
            "Radius (m)",
            "Latitude",
            "Longitude",
            "Reference latitude",
            "Reference longitude",
            "Platform",
            "Browser",
            "Recovered",
            "Coordinates retained",
        ],
        [
            [
                row.get("created_at"),
                row.get("attendance_date"),
                row.get("schedule_label") or "",
                row.get("evidence_snapshot_source") or "unspecified",
                row.get("full_name") or "",
                row.get("university_id") or "",
                row.get("attempt_type"),
                row.get("outcome"),
                LOCATION_REASON_LABELS.get(
                    str(row.get("reason_code")), str(row.get("reason_code"))
                ),
                row.get("accuracy_m"),
                row.get("distance_m"),
                row.get("radius_m"),
                row.get("latitude"),
                row.get("longitude"),
                row.get("reference_latitude"),
                row.get("reference_longitude"),
                row.get("platform"),
                row.get("browser_family"),
                "Yes" if row.get("recovered_at") else "No",
                "Yes" if row.get("latitude") is not None else "No",
            ]
            for row in events
        ],
    )

    calibration_sheet = workbook.create_sheet("Calibration Audit")
    _write_rows(
        calibration_sheet,
        [
            "Time",
            "Manager",
            "Previous latitude",
            "Previous longitude",
            "New latitude",
            "New longitude",
            "Readings",
            "Median accuracy (m)",
        ],
        [
            [
                row["created_at"],
                row["actor_identifier"],
                row["previous_latitude"],
                row["previous_longitude"],
                row["new_latitude"],
                row["new_longitude"],
                row["reading_count"],
                row["median_accuracy_m"],
            ]
            for row in calibrations
        ],
    )

    location_audit_sheet = workbook.create_sheet("Location Rule Audit")
    _write_rows(
        location_audit_sheet,
        [
            "Time",
            "Manager",
            "Change type",
            "Previous latitude",
            "Previous longitude",
            "Previous radius (m)",
            "New latitude",
            "New longitude",
            "New radius (m)",
        ],
        [
            [
                row["created_at"],
                row["actor_identifier"],
                row["change_type"],
                row["previous_latitude"],
                row["previous_longitude"],
                row["previous_radius_m"],
                row["new_latitude"],
                row["new_longitude"],
                row["new_radius_m"],
            ]
            for row in (location_changes or [])
        ],
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_rows(sheet, headers: list[str], rows: list[list]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, header in enumerate(headers, start=1):
        values = [header]
        values.extend(
            str(sheet.cell(row=row, column=column_index).value or "")
            for row in range(2, min(sheet.max_row, 250) + 1)
        )
        width = min(max(len(value) for value in values) + 2, 48)
        sheet.column_dimensions[get_column_letter(column_index)].width = max(width, 10)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
