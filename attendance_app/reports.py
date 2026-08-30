from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from attendance_app.utils import generate_expected_occurrences, weekday_label

REPORT_VERSION = "2.0"
REPORT_SHEETS = [
    "Executive Summary",
    "Course Details",
    "Roster",
    "Timetable",
    "Attendance Records",
    "Student Performance",
    "Lecture Analytics",
    "Security Alerts",
    "Device Audit",
    "OTP Activity",
]

NAVY = "132238"
NAVY_LIGHT = "203752"
TEAL = "00A896"
TEAL_LIGHT = "DDF5F1"
GOLD = "F4B860"
RED = "C44536"
RED_LIGHT = "FCE8E6"
GREEN = "198754"
GREEN_LIGHT = "E5F4EC"
INK = "1C2733"
MUTED = "667085"
LINE = "D8E0E8"
PAPER = "F5F7FA"
WHITE = "FFFFFF"

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=NAVY_LIGHT)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SOFT_FILL = PatternFill("solid", fgColor=PAPER)
TEAL_FILL = PatternFill("solid", fgColor=TEAL_LIGHT)
GOLD_FILL = PatternFill("solid", fgColor="FFF3D9")
THIN_LINE = Side(style="thin", color=LINE)
CARD_BORDER = Border(left=THIN_LINE, right=THIN_LINE, top=THIN_LINE, bottom=THIN_LINE)
BODY_FONT = Font(name="Aptos", color=INK, size=9)
BODY_ALIGNMENT = Alignment(vertical="top")
BODY_WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
ROW_BORDER = Border(bottom=THIN_LINE)


class _Formula(str):
    pass


def build_course_report_xlsx(
    *,
    course,
    students: Iterable,
    schedules: Iterable,
    attendance_records: Iterable,
    eligibility_rows: list[dict[str, object]],
    generated_at: datetime,
    security_alerts: Iterable = (),
    device_audit_events: Iterable = (),
    otp_activity: Iterable = (),
) -> bytes:
    students = list(students)
    schedules = list(schedules)
    attendance_records = list(attendance_records)
    security_alerts = list(security_alerts)
    device_audit_events = list(device_audit_events)
    otp_activity = list(otp_activity)
    lecture_rows = _build_lecture_analytics(
        course=course,
        schedules=schedules,
        attendance_records=attendance_records,
        generated_at=generated_at,
        roster_count=len(students),
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in REPORT_SHEETS}
    workbook.properties.creator = "ClassPresence"
    workbook.properties.title = f"{course['code']} complete attendance report"
    workbook.properties.subject = "Attendance, eligibility, lecture, and security reporting"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    _build_course_details_sheet(sheets["Course Details"], course, generated_at)
    roster_last_row = _build_roster_sheet(sheets["Roster"], students)
    timetable_last_row = _build_timetable_sheet(sheets["Timetable"], schedules)
    attendance_last_row = _build_attendance_sheet(
        sheets["Attendance Records"],
        attendance_records,
    )
    performance_last_row = _build_performance_sheet(
        sheets["Student Performance"],
        eligibility_rows,
    )
    lecture_last_row = _build_lecture_sheet(sheets["Lecture Analytics"], lecture_rows)
    alerts_last_row = _build_security_alerts_sheet(
        sheets["Security Alerts"],
        security_alerts,
    )
    audit_last_row = _build_device_audit_sheet(
        sheets["Device Audit"],
        device_audit_events,
    )
    otp_last_row = _build_otp_activity_sheet(
        sheets["OTP Activity"],
        otp_activity,
        generated_at,
    )
    _build_executive_summary(
        sheets["Executive Summary"],
        course=course,
        generated_at=generated_at,
        roster_last_row=roster_last_row,
        timetable_last_row=timetable_last_row,
        attendance_last_row=attendance_last_row,
        performance_last_row=performance_last_row,
        lecture_last_row=lecture_last_row,
        alerts_last_row=alerts_last_row,
        audit_last_row=audit_last_row,
        otp_last_row=otp_last_row,
        lecture_count=len(lecture_rows),
        performance_count=len(eligibility_rows),
    )

    workbook.active = 0
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_executive_summary(
    sheet,
    *,
    course,
    generated_at: datetime,
    roster_last_row: int,
    timetable_last_row: int,
    attendance_last_row: int,
    performance_last_row: int,
    lecture_last_row: int,
    alerts_last_row: int,
    audit_last_row: int,
    otp_last_row: int,
    lecture_count: int,
    performance_count: int,
) -> None:
    _prepare_sheet(
        sheet,
        title="Executive Summary",
        subtitle=f"{course['code']} | {course['title']}",
        max_column=10,
        freeze_at=None,
    )
    sheet["A3"] = f"Generated {_display_timestamp(generated_at)} | Report version {REPORT_VERSION}"
    sheet["A3"].font = Font(name="Aptos", size=9, color=MUTED)

    cards = [
        ("A4:B4", "A5:B6", "Students", f"=COUNTA('Roster'!$A$6:$A${roster_last_row})", "0"),
        (
            "C4:D4",
            "C5:D6",
            "Attendance Records",
            f"=COUNTA('Attendance Records'!$A$6:$A${attendance_last_row})",
            "0",
        ),
        (
            "E4:F4",
            "E5:F6",
            "Attendance Rate",
            f"=IFERROR(SUM('Student Performance'!$C$6:$C${performance_last_row})/"
            f"SUM('Student Performance'!$E$6:$E${performance_last_row}),0)",
            "0.0%",
        ),
        (
            "G4:H4",
            "G5:H6",
            "Not Eligible",
            f"=COUNTIF('Student Performance'!$I$6:$I${performance_last_row},\"Not eligible\")",
            "0",
        ),
        (
            "I4:J4",
            "I5:J6",
            "Open Alerts",
            f"=COUNTIF('Security Alerts'!$K$6:$K${alerts_last_row},\"Open\")",
            "0",
        ),
        (
            "A8:B8",
            "A9:B10",
            "Weekly Windows",
            f"=COUNTA('Timetable'!$A$6:$A${timetable_last_row})",
            "0",
        ),
        (
            "C8:D8",
            "C9:D10",
            "Elapsed Lectures",
            f"=COUNTA('Lecture Analytics'!$A$6:$A${lecture_last_row})",
            "0",
        ),
        (
            "E8:F8",
            "E9:F10",
            "Registered Devices",
            f"=COUNTIF('Roster'!$E$6:$E${roster_last_row},\"Registered\")",
            "0",
        ),
        (
            "G8:H8",
            "G9:H10",
            "Device Events",
            f"=COUNTA('Device Audit'!$A$6:$A${audit_last_row})",
            "0",
        ),
        (
            "I8:J8",
            "I9:J10",
            "OTP Events",
            f"=COUNTA('OTP Activity'!$A$6:$A${otp_last_row})",
            "0",
        ),
    ]
    for label_range, value_range, label, formula, number_format in cards:
        sheet.merge_cells(label_range)
        sheet.merge_cells(value_range)
        label_cell = sheet[label_range.split(":")[0]]
        value_cell = sheet[value_range.split(":")[0]]
        label_cell.value = label
        value_cell.value = formula
        label_cell.fill = SECTION_FILL
        label_cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.fill = PatternFill("solid", fgColor=WHITE)
        value_cell.font = Font(name="Aptos Display", size=20, bold=True, color=INK)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format
        for row in sheet[label_range]:
            for cell in row:
                cell.border = CARD_BORDER
        for row in sheet[value_range]:
            for cell in row:
                cell.border = CARD_BORDER

    sheet["A33"] = "Eligibility"
    sheet["B33"] = "Students"
    sheet["A34"] = "Eligible"
    sheet["B34"] = f"=COUNTIF('Student Performance'!$I$6:$I${performance_last_row},\"Eligible\")"
    sheet["A35"] = "Not eligible"
    sheet["B35"] = (
        f"=COUNTIF('Student Performance'!$I$6:$I${performance_last_row},\"Not eligible\")"
    )
    for cell in sheet[33]:
        if cell.column > 2:
            continue
        cell.fill = HEADER_FILL
        cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    for row_number in (34, 35):
        for cell in sheet[row_number][:2]:
            cell.border = CARD_BORDER
            cell.font = Font(name="Aptos", size=9, color=INK)

    if performance_count:
        eligibility_chart = DoughnutChart()
        eligibility_chart.title = "Eligibility distribution"
        eligibility_chart.height = 7.2
        eligibility_chart.width = 11.5
        eligibility_chart.holeSize = 62
        eligibility_chart.legend.position = "b"
        eligibility_chart.add_data(
            Reference(sheet, min_col=2, min_row=33, max_row=35),
            titles_from_data=True,
        )
        eligibility_chart.set_categories(Reference(sheet, min_col=1, min_row=34, max_row=35))
        eligibility_chart.dataLabels = DataLabelList()
        eligibility_chart.dataLabels.showPercent = True
        eligibility_chart.dataLabels.showLeaderLines = True
        sheet.add_chart(eligibility_chart, "A13")

    if lecture_count:
        chart_last_row = min(lecture_last_row, 20)
        attendance_chart = BarChart()
        attendance_chart.type = "col"
        attendance_chart.style = 10
        attendance_chart.title = "Attendance rate by elapsed lecture"
        attendance_chart.y_axis.title = "Attendance rate"
        attendance_chart.y_axis.numFmt = "0%"
        attendance_chart.y_axis.scaling.min = 0
        attendance_chart.y_axis.scaling.max = 1
        attendance_chart.x_axis.title = "Lecture date"
        attendance_chart.height = 7.2
        attendance_chart.width = 14.5
        attendance_chart.legend = None
        attendance_chart.add_data(
            Reference(
                sheet.parent["Lecture Analytics"],
                min_col=9,
                min_row=5,
                max_row=chart_last_row,
            ),
            titles_from_data=True,
        )
        attendance_chart.set_categories(
            Reference(
                sheet.parent["Lecture Analytics"],
                min_col=2,
                min_row=6,
                max_row=chart_last_row,
            )
        )
        sheet.add_chart(attendance_chart, "F13")

    sheet["A29"] = "Reporting scope"
    sheet["A29"].font = Font(name="Aptos", size=10, bold=True, color=INK)
    sheet.merge_cells("A30:J31")
    sheet["A30"] = (
        "This workbook contains complete course reporting records. OTP values, OTP hashes, "
        "passkey keys, credential IDs, and raw device-binding hashes are intentionally excluded."
    )
    sheet["A30"].fill = GOLD_FILL
    sheet["A30"].font = Font(name="Aptos", size=9, color=INK)
    sheet["A30"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet["A30"].border = CARD_BORDER
    _set_widths(sheet, [14] * 10)
    sheet.sheet_view.zoomScale = 85


def _build_course_details_sheet(sheet, course, generated_at: datetime) -> None:
    _prepare_sheet(
        sheet,
        title="Course Details",
        subtitle="Configuration and reporting scope",
        max_column=4,
        freeze_at=None,
    )
    items = [
        ("Report Version", REPORT_VERSION),
        ("Course Code", course["code"]),
        ("Course Name", course["title"]),
        ("Start Date", _excel_date(course["start_date"])),
        ("End Date", _excel_date(course["end_date"] or course["start_date"])),
        ("Total Meetings", int(course["total_meetings"])),
        ("Latitude", float(course["latitude"])),
        ("Longitude", float(course["longitude"])),
        ("Allowed Radius (m)", float(course["radius_m"])),
        ("Absence Limit (%)", float(course["absence_limit_pct"])),
        ("Generated At", _excel_datetime(generated_at)),
        ("Timezone", _timezone_name(generated_at)),
        (
            "Security Scope",
            "Report-safe metadata only; passkey, credential, OTP, and raw device secrets excluded.",
        ),
    ]
    start_row = 5
    for row_index, (label, value) in enumerate(items, start=start_row):
        sheet.cell(row=row_index, column=1, value=label)
        _write_cell_value(sheet.cell(row=row_index, column=2), value)
        sheet.cell(row=row_index, column=1).fill = SECTION_FILL
        sheet.cell(row=row_index, column=1).font = Font(name="Aptos", bold=True, color=WHITE)
        sheet.cell(row=row_index, column=2).fill = PatternFill("solid", fgColor=WHITE)
        sheet.cell(row=row_index, column=2).font = Font(name="Aptos", color=INK)
        sheet.cell(row=row_index, column=1).border = CARD_BORDER
        sheet.cell(row=row_index, column=2).border = CARD_BORDER
        sheet.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    sheet["B8"].number_format = "yyyy-mm-dd"
    sheet["B9"].number_format = "yyyy-mm-dd"
    sheet["B11"].number_format = "0.000000"
    sheet["B12"].number_format = "0.000000"
    sheet["B13"].number_format = "0.0"
    sheet["B14"].number_format = "0.0"
    sheet["B15"].number_format = "yyyy-mm-dd hh:mm"
    _set_widths(sheet, [25, 75, 3, 3])


def _build_roster_sheet(sheet, students: list) -> int:
    headers = [
        "Student ID",
        "Student Name",
        "Email",
        "Phone",
        "Device Status",
        "Device Reference",
        "Registered At",
        "Last Verified",
        "Device Type",
        "Backup State",
    ]
    rows = []
    for student in students:
        device_id = student.get("registered_device_id")
        rows.append(
            [
                str(student["university_id"]),
                student["full_name"],
                student.get("email") or "",
                _report_phone(student.get("phone")),
                "Registered" if device_id else "Not registered",
                _device_reference(device_id),
                _excel_datetime(student.get("device_registered_at")),
                _excel_datetime(student.get("device_last_used_at")),
                str(student.get("device_type") or "").replace("_", " ").title(),
                "Backed up" if student.get("device_backed_up") else "Not backed up",
            ]
        )
    last_row = _build_table_sheet(
        sheet,
        title="Roster",
        subtitle="Enrolled students and current report-safe device status",
        headers=headers,
        rows=rows,
        table_name="RosterTable",
        widths=[18, 28, 32, 18, 18, 20, 20, 20, 18, 18],
    )
    _format_datetime_columns(sheet, last_row, [7, 8])
    _format_text_columns(sheet, last_row, [1, 4, 6])
    if last_row >= 6:
        sheet.conditional_formatting.add(
            f"E6:E{last_row}",
            FormulaRule(
                formula=['E6="Registered"'], fill=PatternFill("solid", fgColor=GREEN_LIGHT)
            ),
        )
    return last_row


def _build_timetable_sheet(sheet, schedules: list) -> int:
    return _build_table_sheet(
        sheet,
        title="Timetable",
        subtitle="Weekly lecture windows controlling student access",
        headers=["Weekday", "Window Label", "Start Time", "End Time"],
        rows=[
            [
                weekday_label(int(schedule["weekday"])),
                schedule["label"],
                schedule["start_time"],
                schedule["end_time"],
            ]
            for schedule in schedules
        ],
        table_name="TimetableTable",
        widths=[18, 30, 16, 16],
    )


def _build_attendance_sheet(sheet, attendance_records: list) -> int:
    headers = [
        "Record ID",
        "Student Name",
        "Student ID",
        "Date",
        "Lecture",
        "Scheduled Start",
        "Scheduled End",
        "Checked In",
        "Latitude",
        "Longitude",
        "Distance (m)",
        "GPS Accuracy (m)",
        "Verification",
        "Device Reference",
    ]
    rows = []
    for row_index, row in enumerate(attendance_records, start=1):
        registered_device_id = row.get("registered_device_id")
        rows.append(
            [
                int(row.get("attendance_id") or row_index),
                row["full_name"],
                str(row["university_id"]),
                _excel_date(row["attendance_date"]),
                row["schedule_label"],
                row.get("schedule_start_time") or "",
                row.get("schedule_end_time") or "",
                _excel_datetime(row["stamped_at"]),
                _optional_float(row.get("student_latitude")),
                _optional_float(row.get("student_longitude")),
                float(row["distance_m"]),
                _optional_float(row.get("accuracy_m")),
                "Passkey verified"
                if registered_device_id or row.get("device_binding_hash")
                else "Imported / legacy",
                _device_reference(registered_device_id, row.get("device_binding_hash")),
            ]
        )
    last_row = _build_table_sheet(
        sheet,
        title="Attendance Records",
        subtitle="Complete location and device-verification evidence for every check-in",
        headers=headers,
        rows=rows,
        table_name="AttendanceRecordsTable",
        widths=[12, 28, 18, 14, 24, 16, 16, 22, 15, 15, 16, 18, 22, 20],
    )
    _format_date_columns(sheet, last_row, [4])
    _format_datetime_columns(sheet, last_row, [8])
    _format_decimal_columns(sheet, last_row, [9, 10], "0.000000")
    _format_decimal_columns(sheet, last_row, [11, 12], "0.0")
    _format_text_columns(sheet, last_row, [3, 14])
    return last_row


def _build_performance_sheet(sheet, eligibility_rows: list[dict[str, object]]) -> int:
    headers = [
        "Student",
        "University ID",
        "Attended",
        "Absences",
        "Elapsed Meetings",
        "Total Meetings",
        "Attendance Rate",
        "Threshold",
        "Status",
    ]
    rows = []
    for row_number, row in enumerate(eligibility_rows, start=6):
        rows.append(
            [
                row["Student"],
                str(row["University ID"]),
                int(row["Attended"]),
                _Formula(f"=MAX(E{row_number}-C{row_number},0)"),
                int(row["Elapsed Meetings"]),
                int(row["Total Meetings"]),
                _Formula(f"=IFERROR(C{row_number}/F{row_number},0)"),
                _Formula(f"=ROUNDUP(F{row_number}*'Course Details'!$B$14/100,0)"),
                _Formula(
                    f"=IF(AND(H{row_number}>0,D{row_number}>=H{row_number}),"
                    '"Not eligible","Eligible")'
                ),
            ]
        )
    last_row = _build_table_sheet(
        sheet,
        title="Student Performance",
        subtitle="Formula-driven attendance standing and exam eligibility",
        headers=headers,
        rows=rows,
        table_name="StudentPerformanceTable",
        widths=[28, 18, 14, 14, 20, 18, 18, 14, 18],
    )
    _format_integer_columns(sheet, last_row, [3, 4, 5, 6, 8])
    _format_decimal_columns(sheet, last_row, [7], "0.0%")
    _format_text_columns(sheet, last_row, [2])
    if last_row >= 6:
        sheet.conditional_formatting.add(
            f"G6:G{last_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="FCE8E6",
                mid_type="num",
                mid_value=0.75,
                mid_color="FFF3D9",
                end_type="num",
                end_value=1,
                end_color="E5F4EC",
            ),
        )
        sheet.conditional_formatting.add(
            f"I6:I{last_row}",
            FormulaRule(
                formula=['I6="Not eligible"'], fill=PatternFill("solid", fgColor=RED_LIGHT)
            ),
        )
        sheet.conditional_formatting.add(
            f"I6:I{last_row}",
            FormulaRule(formula=['I6="Eligible"'], fill=PatternFill("solid", fgColor=GREEN_LIGHT)),
        )
    return last_row


def _build_lecture_sheet(sheet, lecture_rows: list[dict[str, object]]) -> int:
    headers = [
        "Date",
        "Lecture Reference",
        "Weekday",
        "Lecture",
        "Start",
        "End",
        "Checked In",
        "Missing",
        "Attendance Rate",
        "Status",
    ]
    roster_last_row = max(6, sheet.parent["Roster"].max_row)
    rows = []
    for row_number, row in enumerate(lecture_rows, start=6):
        rows.append(
            [
                row["date"],
                f"{row['date'].isoformat()} | {row['label']}",
                row["weekday"],
                row["label"],
                row["start_time"],
                row["end_time"],
                row["checked_in"],
                _Formula(f"=MAX(0,COUNTA('Roster'!$A$6:$A${roster_last_row})-G{row_number})"),
                _Formula(f"=IFERROR(G{row_number}/COUNTA('Roster'!$A$6:$A${roster_last_row}),0)"),
                row["status"],
            ]
        )
    last_row = _build_table_sheet(
        sheet,
        title="Lecture Analytics",
        subtitle="Attendance totals and rates for every elapsed lecture",
        headers=headers,
        rows=rows,
        table_name="LectureAnalyticsTable",
        widths=[14, 38, 16, 28, 14, 14, 16, 14, 20, 16],
    )
    _format_date_columns(sheet, last_row, [1])
    _format_integer_columns(sheet, last_row, [7, 8])
    _format_decimal_columns(sheet, last_row, [9], "0.0%")
    if last_row >= 6:
        sheet.conditional_formatting.add(
            f"I6:I{last_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="FCE8E6",
                mid_type="num",
                mid_value=0.75,
                mid_color="FFF3D9",
                end_type="num",
                end_value=1,
                end_color="E5F4EC",
            ),
        )
    return last_row


def _build_security_alerts_sheet(sheet, security_alerts: list) -> int:
    headers = [
        "Alert ID",
        "Created At",
        "Date",
        "Student",
        "Student ID",
        "Lecture",
        "Severity",
        "Event",
        "Details",
        "Device Reference",
        "Status",
        "Resolved At",
        "Latitude",
        "Longitude",
        "GPS Accuracy (m)",
    ]
    rows = []
    for row in security_alerts:
        rows.append(
            [
                f"ALT-{int(row['id']):06d}",
                _excel_datetime(row["created_at"]),
                _excel_date(row.get("attendance_date")),
                row.get("full_name") or "Unknown",
                str(row.get("university_id") or ""),
                row.get("schedule_label") or "",
                str(row["severity"]).title(),
                str(row["alert_type"]).replace("_", " ").title(),
                row["message"],
                _device_reference(None, row.get("device_binding_hash")),
                "Resolved" if row.get("resolved_at") else "Open",
                _excel_datetime(row.get("resolved_at")),
                _optional_float(row.get("latitude")),
                _optional_float(row.get("longitude")),
                _optional_float(row.get("accuracy_m")),
            ]
        )
    last_row = _build_table_sheet(
        sheet,
        title="Security Alerts",
        subtitle="Proxy indicators, severity, evidence, and manager resolution status",
        headers=headers,
        rows=rows,
        table_name="SecurityAlertsTable",
        widths=[16, 22, 14, 26, 18, 24, 14, 30, 55, 20, 14, 22, 15, 15, 18],
    )
    _format_datetime_columns(sheet, last_row, [2, 12])
    _format_date_columns(sheet, last_row, [3])
    _format_decimal_columns(sheet, last_row, [13, 14], "0.000000")
    _format_decimal_columns(sheet, last_row, [15], "0.0")
    _format_text_columns(sheet, last_row, [1, 5, 10])
    if last_row >= 6:
        sheet.conditional_formatting.add(
            f"G6:G{last_row}",
            FormulaRule(formula=['G6="Critical"'], fill=PatternFill("solid", fgColor=RED_LIGHT)),
        )
        sheet.conditional_formatting.add(
            f"K6:K{last_row}",
            FormulaRule(formula=['K6="Open"'], fill=PatternFill("solid", fgColor="FFF3D9")),
        )
    return last_row


def _build_device_audit_sheet(sheet, events: list) -> int:
    headers = [
        "Event ID",
        "Created At",
        "Student",
        "Student ID",
        "Event",
        "Actor Type",
        "Actor",
        "Course",
        "Previous Device",
        "New Device",
    ]
    rows = [
        [
            f"DEV-EVT-{int(row['id']):06d}",
            _excel_datetime(row["created_at"]),
            row["student_name"],
            str(row["university_id"]),
            str(row["event_type"]).replace("_", " ").title(),
            str(row["actor_type"]).title(),
            row["actor_identifier"],
            row.get("course_code") or "",
            _device_reference(row.get("previous_device_id")),
            _device_reference(row.get("new_device_id")),
        ]
        for row in events
    ]
    last_row = _build_table_sheet(
        sheet,
        title="Device Audit",
        subtitle="Permanent registration and manager reset history",
        headers=headers,
        rows=rows,
        table_name="DeviceAuditTable",
        widths=[20, 22, 28, 18, 30, 16, 24, 16, 20, 20],
    )
    _format_datetime_columns(sheet, last_row, [2])
    _format_text_columns(sheet, last_row, [1, 4, 7, 8, 9, 10])
    return last_row


def _build_otp_activity_sheet(sheet, otp_activity: list, generated_at: datetime) -> int:
    headers = [
        "OTP Reference",
        "Created At",
        "Date",
        "Student",
        "Student ID",
        "Lecture",
        "Delivery",
        "Target",
        "Expires At",
        "Status",
        "Used At",
        "Invalidated At",
    ]
    rows = [
        [
            f"OTP-{int(row['id']):06d}",
            _excel_datetime(row["created_at"]),
            _excel_date(row.get("attendance_date")),
            row["full_name"],
            str(row["university_id"]),
            row.get("schedule_label") or "",
            str(row["delivery_method"]).title(),
            _mask_delivery_target(row.get("delivery_target")),
            _excel_datetime(row["expires_at"]),
            _otp_status(row, generated_at),
            _excel_datetime(row.get("used_at")),
            _excel_datetime(row.get("invalidated_at")),
        ]
        for row in otp_activity
    ]
    last_row = _build_table_sheet(
        sheet,
        title="OTP Activity",
        subtitle="Lecture-bound issuance lifecycle without OTP values or hashes",
        headers=headers,
        rows=rows,
        table_name="OtpActivityTable",
        widths=[18, 22, 14, 28, 18, 24, 14, 28, 22, 16, 22, 22],
    )
    _format_datetime_columns(sheet, last_row, [2, 9, 11, 12])
    _format_date_columns(sheet, last_row, [3])
    _format_text_columns(sheet, last_row, [1, 5, 8])
    return last_row


def _build_lecture_analytics(
    *,
    course,
    schedules: list,
    attendance_records: list,
    generated_at: datetime,
    roster_count: int,
) -> list[dict[str, object]]:
    counts: Counter[tuple[int, str]] = Counter()
    fallback_counts: Counter[tuple[str, str]] = Counter()
    for row in attendance_records:
        attendance_date = str(row["attendance_date"])
        if row.get("schedule_id") is not None:
            counts[(int(row["schedule_id"]), attendance_date)] += 1
        fallback_counts[(str(row["schedule_label"]), attendance_date)] += 1

    occurrences = generate_expected_occurrences(
        course["start_date"],
        course["end_date"] or course["start_date"],
        schedules,
        generated_at,
        only_elapsed=True,
    )
    result = []
    seen_keys: set[tuple[str, str]] = set()
    for occurrence in occurrences:
        attendance_date = occurrence.starts_at.date().isoformat()
        fallback_key = (occurrence.label, attendance_date)
        checked_in = counts.get(
            (occurrence.schedule_id, attendance_date),
            fallback_counts.get(fallback_key, 0),
        )
        seen_keys.add(fallback_key)
        result.append(
            {
                "date": occurrence.starts_at.date(),
                "weekday": occurrence.starts_at.strftime("%A"),
                "label": occurrence.label,
                "start_time": occurrence.starts_at.strftime("%H:%M"),
                "end_time": occurrence.ends_at.strftime("%H:%M"),
                "checked_in": checked_in,
                "missing": max(roster_count - checked_in, 0),
                "attendance_rate": checked_in / roster_count if roster_count else 0.0,
                "status": "Complete",
            }
        )

    for (label, attendance_date), checked_in in fallback_counts.items():
        if (label, attendance_date) in seen_keys:
            continue
        parsed_date = _excel_date(attendance_date)
        result.append(
            {
                "date": parsed_date,
                "weekday": parsed_date.strftime("%A") if parsed_date else "",
                "label": label,
                "start_time": "",
                "end_time": "",
                "checked_in": checked_in,
                "missing": max(roster_count - checked_in, 0),
                "attendance_rate": checked_in / roster_count if roster_count else 0.0,
                "status": "Recorded",
            }
        )
    return sorted(result, key=lambda row: (row["date"] or date.min, str(row["start_time"])))


def _build_table_sheet(
    sheet,
    *,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[list[object]],
    table_name: str,
    widths: list[float],
) -> int:
    _prepare_sheet(
        sheet,
        title=title,
        subtitle=subtitle,
        max_column=len(headers),
        freeze_at="A6",
    )
    header_row = 5
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = Font(name="Aptos", color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=TEAL))
    sheet.row_dimensions[header_row].height = 28

    for row_index, row in enumerate(rows, start=header_row + 1):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            _write_cell_value(cell, value)
            cell.font = BODY_FONT
            cell.alignment = BODY_WRAP_ALIGNMENT if column_index in {2, 3, 8, 9} else BODY_ALIGNMENT
            cell.border = ROW_BORDER
        sheet.row_dimensions[row_index].height = 20

    last_row = header_row + len(rows)
    formula_last_row = max(last_row, header_row + 1)
    if rows:
        table = Table(
            displayName=table_name, ref=f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        sheet.auto_filter.ref = table.ref
    else:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
        sheet.merge_cells(
            start_row=3,
            start_column=1,
            end_row=3,
            end_column=min(len(headers), 6),
        )
        note = sheet.cell(row=3, column=1, value="No records available for this report section.")
        note.font = Font(name="Aptos", italic=True, color=MUTED)
        note.fill = SOFT_FILL
        note.alignment = Alignment(vertical="center")
    _set_widths(sheet, widths)
    sheet.sheet_view.zoomScale = 90
    return formula_last_row


def _prepare_sheet(
    sheet,
    *,
    title: str,
    subtitle: str,
    max_column: int,
    freeze_at: str | None,
) -> None:
    title_width = max(max_column, 4)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_width)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=title_width)
    sheet["A1"] = title
    _write_cell_value(sheet["A2"], subtitle)
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet["A2"].fill = TITLE_FILL
    sheet["A2"].font = Font(name="Aptos", size=10, color="C8D6E5")
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.row_dimensions[2].height = 22
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = freeze_at
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:5"
    sheet.sheet_properties.outlinePr.summaryBelow = True


def _set_widths(sheet, widths: list[float]) -> None:
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width, 10), 60)


def _format_date_columns(sheet, last_row: int, columns: list[int]) -> None:
    for column in columns:
        for row in range(6, last_row + 1):
            sheet.cell(row=row, column=column).number_format = "yyyy-mm-dd"


def _format_datetime_columns(sheet, last_row: int, columns: list[int]) -> None:
    for column in columns:
        for row in range(6, last_row + 1):
            sheet.cell(row=row, column=column).number_format = "yyyy-mm-dd hh:mm"


def _format_decimal_columns(sheet, last_row: int, columns: list[int], number_format: str) -> None:
    for column in columns:
        for row in range(6, last_row + 1):
            sheet.cell(row=row, column=column).number_format = number_format


def _format_integer_columns(sheet, last_row: int, columns: list[int]) -> None:
    _format_decimal_columns(sheet, last_row, columns, "0")


def _format_text_columns(sheet, last_row: int, columns: list[int]) -> None:
    for column in columns:
        for row in range(6, last_row + 1):
            sheet.cell(row=row, column=column).number_format = "@"


def _write_cell_value(cell, value) -> None:
    if isinstance(value, _Formula):
        cell.value = str(value)
        return
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        cell.value = f"'{value}"
        cell.quotePrefix = True
        return
    cell.value = value


def _excel_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value)[:10])


def _excel_datetime(value) -> datetime | None:
    if value in (None, ""):
        return None
    parsed = value if isinstance(value, datetime) else datetime.fromisoformat(str(value))
    return parsed.replace(tzinfo=None)


def _display_timestamp(value: datetime) -> str:
    return value.strftime("%Y-%m-%d %H:%M") + f" {_timezone_name(value)}"


def _timezone_name(value: datetime) -> str:
    return str(getattr(value.tzinfo, "key", None) or value.tzname() or "Local")


def _optional_float(value) -> float | None:
    return None if value in (None, "") else float(value)


def _device_reference(device_id, device_binding_hash=None) -> str:
    if device_id not in (None, ""):
        return f"DEV-{int(device_id):06d}"
    if device_binding_hash:
        digest = sha256(str(device_binding_hash).encode("utf-8")).hexdigest()[:8].upper()
        return f"DEV-{digest}"
    return ""


def _mask_delivery_target(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "@" in text:
        local, domain = text.split("@", 1)
        visible = local[:1]
        return f"{visible}{'*' * max(len(local) - 1, 3)}@{domain}"
    if len(text) <= 4:
        return "*" * len(text)
    return f"{'*' * (len(text) - 4)}{text[-4:]}"


def _report_phone(value) -> str:
    text = str(value or "").strip()
    if text.startswith("+"):
        return f"tel:{text}"
    return text


def _otp_status(row, generated_at: datetime) -> str:
    if row.get("used_at"):
        return "Used"
    if row.get("invalidated_at"):
        return "Invalidated"
    expires_at = row.get("expires_at")
    if expires_at:
        expiry = (
            expires_at
            if isinstance(expires_at, datetime)
            else datetime.fromisoformat(str(expires_at))
        )
        comparison_time = generated_at
        if expiry.tzinfo is None and comparison_time.tzinfo is not None:
            comparison_time = comparison_time.replace(tzinfo=None)
        elif expiry.tzinfo is not None and comparison_time.tzinfo is None:
            expiry = expiry.replace(tzinfo=None)
        if expiry <= comparison_time:
            return "Expired"
    return "Active"
